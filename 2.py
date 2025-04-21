import re
import openpyxl

# Aadhaar validation helpers (Verhoeff algorithm)
mult = [[0, 1, 2, 3, 4, 5, 6, 7, 8, 9], [1, 2, 3, 4, 0, 6, 7, 8, 9, 5],
        [2, 3, 4, 0, 1, 7, 8, 9, 5, 6], [3, 4, 0, 1, 2, 8, 9, 5, 6, 7],
        [4, 0, 1, 2, 3, 9, 5, 6, 7, 8], [5, 9, 8, 7, 6, 0, 4, 3, 2, 1],
        [6, 5, 9, 8, 7, 1, 0, 4, 3, 2], [7, 6, 5, 9, 8, 2, 1, 0, 4, 3],
        [8, 7, 6, 5, 9, 3, 2, 1, 0, 4], [9, 8, 7, 6, 5, 4, 3, 2, 1, 0]]

perm = [[0, 1, 2, 3, 4, 5, 6, 7, 8, 9], [1, 5, 7, 6, 2, 8, 3, 0, 9, 4],
        [5, 8, 0, 3, 7, 9, 6, 1, 4, 2], [8, 9, 1, 6, 0, 4, 3, 5, 2, 7],
        [9, 4, 5, 3, 1, 2, 6, 8, 7, 0], [4, 2, 8, 6, 5, 7, 3, 9, 0, 1],
        [2, 7, 9, 3, 8, 0, 6, 4, 1, 5], [7, 0, 4, 6, 9, 1, 3, 2, 5, 8]]

def Validate(aadharNum):
    try:
        i = len(aadharNum)
        j = 0
        x = 0
        while i > 0:
            i -= 1
            x = mult[x][perm[(j % 8)][int(aadharNum[i])]]
            j += 1
        return 'Valid' if x == 0 else 'Invalid'
    except (ValueError, IndexError):
        return 'Invalid'

# File names
input_file = 'input.txt'
output_file = 'output.xlsx'

# Create Excel workbook and worksheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = '12 Digit Numbers'

# Headers
sheet['A1'] = '12-Digit Number'
sheet['B1'] = 'Line Found'
sheet['C1'] = 'Validation'

# Regex pattern: capture exactly 12-digit sequences not part of longer numbers
pattern = r'(?<!\d)(\d{12})(?!\d)'

# Count total lines
with open(input_file, 'r', encoding='utf-8') as f:
    total_lines = sum(1 for _ in f)

# Now process the file
row = 2
line_count = 0

with open(input_file, 'r', encoding='utf-8') as file:
    for line in file:
        line_count += 1
        matches = re.findall(pattern, line)
        for number in matches:
            sheet.cell(row=row, column=1).value = number
            sheet.cell(row=row, column=2).value = line.strip()
            sheet.cell(row=row, column=3).value = Validate(number)
            row += 1

# Save Excel
workbook.save(output_file)

# Print summary
print(f"Total lines in file: {total_lines}")
print(f"Lines analyzed: {line_count}")
print(f"Done! Output saved to '{output_file}'.")
